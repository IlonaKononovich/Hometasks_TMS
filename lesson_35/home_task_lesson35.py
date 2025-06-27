# Импортируем основные компоненты Airflow для создания DAG и задач
from airflow import DAG
from airflow.providers.postgres.hooks.postgres import PostgresHook      # Хук для работы с Postgres
from airflow.operators.dummy import DummyOperator                       # Заглушка — простой оператор без действия
from airflow.operators.python import PythonOperator                     # Оператор для запуска Python-функций
from datetime import datetime                                           # Работа с датами
from openpyxl import load_workbook                                      # Чтение Excel-файлов


# Задаём базовые параметры для DAG, например, владелец задач
default_args = {
    'owner': "ilona",
}

# Функция для чтения данных из Excel и загрузки в Postgres с обработкой конфликтов
def read_xlsx_and_load_to_postgres(
    filepath: str,                                                      # Путь к Excel-файлу
    table_name: str,                                                    # Имя таблицы для загрузки в БД
    columns: list,                                                      # Список колонок для вставки
    conflict_key: str = 'id',                                           # Колонка для проверки конфликтов (уникальный ключ)
    postgres_conn_id: str = 'my_postgres_conn'                          # Идентификатор подключения к Postgres в Airflow
):

    # Загружаем Excel-файл
    wb = load_workbook(filename=filepath)
    # Берём активный лист (обычно первый)
    ws = wb.active

    # Получаем все строки в виде списка значений (без форматирования)
    rows = list(ws.iter_rows(values_only=True))
    # Отбрасываем заголовок — первая строка с названиями колонок
    data_rows = rows[1:]

    # Выводим начало загрузки данных
    print(f"[Загрузка из {filepath}]")

    # Создаём подключение к Postgres через Airflow Hook
    pg_hook = PostgresHook(postgres_conn_id=postgres_conn_id)
    # Получаем объект подключения
    conn = pg_hook.get_conn() 
    # Получаем курсор для выполнения SQL-запросов
    cursor = conn.cursor()    

    # Формируем строку с именами колонок через запятую для SQL-запроса
    columns_str = ', '.join(columns)
    # Формируем плейсхолдеры %s для безопасной вставки значений (по количеству колонок)
    placeholders = ', '.join(['%s'] * len(columns))
    # Формируем SQL-запрос для вставки данных с игнорированием конфликтов по ключу
    insert_sql = f"""
        INSERT INTO {table_name} ({columns_str})
        VALUES ({placeholders})
        ON CONFLICT ({conflict_key}) DO NOTHING
    """

    # Цикл по всем строкам из Excel (без заголовка)
    for row in data_rows:
        # Выполняем SQL вставку с подстановкой значений из текущей строки
        cursor.execute(insert_sql, row)

    # Фиксируем изменения в базе (commit)
    conn.commit()
    # Закрываем курсор и соединение
    cursor.close()
    conn.close()

    # Выводим успешную вставку данных
    print(f"[Данные вставились в {table_name}]")

# Функция, запускающая SQL-запрос для обновления данных в data mart
def run_sql_to_data_mart():

    # Создаём подключение к Postgres через Hook
    pg_hook = PostgresHook(postgres_conn_id='my_postgres_conn')
    # SQL для вставки/обновления агрегированных данных
    sql = """
    -- Вставляем в таблицу итогов по пользователям
    INSERT INTO data_mart.user_orders_summary (name, card_number, total_order_sum)  
    SELECT
        u.name,                                                            -- Имя пользователя из таблицы пользователей
        u.card_number,                                                     -- Номер карты пользователя (уникальный идентификатор)
        COALESCE(SUM(o.total_cost), 0)                                     -- Суммируем стоимость всех заказов пользователя, если заказов нет — 0
        AS total_order_sum
    FROM raw.users u                                                       -- Основная таблица пользователей
    LEFT JOIN raw.orders o                                                 -- Левое соединение с таблицей заказов (берём всех пользователей, даже без заказов)
      ON u.card_number = o.card_number                                     -- Связь по номеру карты
    GROUP BY u.name, u.card_number                                         -- Группируем по пользователю для агрегации сумм заказов
    ON CONFLICT (card_number) DO UPDATE                                    -- При конфликте по card_number (если запись уже есть)
    SET total_order_sum = EXCLUDED.total_order_sum,                        -- Обновляем сумму заказов на новую
        name = EXCLUDED.name;                                              -- Обновляем имя (на всякий случай)
    """

    # Выводим начало обновления data mart
    print("Запуск обновления data_mart.user_orders_summary")
    # Запускаем SQL-запрос через Hook
    pg_hook.run(sql)
    # Выводим окончание обновления
    print("Обновление data_mart.user_orders_summary завершено")

# Определяем DAG (граф задач) — как и когда будет запускаться процесс
with DAG(
    'import_data_from_excel',                                               # Имя DAG
    default_args=default_args,                                              # Базовые параметры
    description='Интеграция данных из xlsx',                                # Описание DAG
    schedule_interval=None,                                                 # Запуск только вручную (нет расписания)
    start_date=datetime(2025, 6, 1),                                        # Дата старта DAG (важно для Airflow)
    catchup=False                                                           # Не запускать пропущенные запуски при старте DAG
) as dag:
    
    # Заглушка — удобна для упрощения цепочки зависимостей
    d = DummyOperator(task_id="dummy")

    # Задача на чтение и загрузку пользователей
    read_and_load_users_task = PythonOperator(
        task_id='read_and_load_users',                                      # Уникальный id задачи
        python_callable=read_xlsx_and_load_to_postgres,                     # Вызываемая функция
        op_kwargs={                                                         # Аргументы функции
            'filepath': '/opt/airflow/dags/data/users_data.xlsx',
            'table_name': 'raw.users',
            'columns': ['name', 'surname', 'city', 'age', 'card_number'],
            'conflict_key': 'card_number'                                   # Ключ для обработки конфликтов
        }
    )

    # Задача на чтение и загрузку заказов
    read_and_load_orders_task = PythonOperator(
        task_id='read_and_load_orders',
        python_callable=read_xlsx_and_load_to_postgres,
        op_kwargs={
            'filepath': '/opt/airflow/dags/data/orders_data.xlsx',
            'table_name': 'raw.orders',
            'columns': ['id', 'name', 'count', 'price', 'total_cost', 'card_number'],
            'conflict_key': 'id'
        }
    )
    
    # Задача на чтение и загрузку данных о доставках
    read_and_load_deliveries_data_task = PythonOperator(
        task_id='read_and_load_deliveries_data',
        python_callable=read_xlsx_and_load_to_postgres,
        op_kwargs={
            'filepath': '/opt/airflow/dags/data/deliveries_data.xlsx',
            'table_name': 'raw.deliveries_data',
            'columns': [
                'delivery_number',
                'order_id',
                'product_array',
                'delivery_company',
                'delivery_cost',
                'courier_name',
                'courier_phone',
                'start_time',
                'end_time',
                'delivery_city',
                'warehouse_from',
                'warehouse_address'
            ],
            'conflict_key': 'delivery_number'
        }
    )

    # Задача на обновление data mart через SQL-запрос
    run_data_mart_task = PythonOperator(
        task_id='run_data_mart_update',
        python_callable=run_sql_to_data_mart,
    )

    # Определяем порядок выполнения задач — последовательно
    d >> read_and_load_users_task >> read_and_load_orders_task >> read_and_load_deliveries_data_task >> run_data_mart_task
